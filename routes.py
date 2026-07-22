import urllib.request
import json as json_lib
import random
import smtplib
import os
from email.mime.text import MIMEText
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from database import get_db, User, Submission, hash_password, verify_password
from pydantic import BaseModel
from datetime import datetime
from typing import Optional

router = APIRouter()

# ── INPUT MODELS ──
class SignupRequest(BaseModel):
    name: str
    email: str
    password: str
    role: str  # "manager" or "investment_manager"
    asset_class: Optional[str] = None
    institution_name: Optional[str] = None
    institution_location: Optional[str] = None
    job_title: Optional[str] = None
    phone_number: Optional[str] = None

class LoginRequest(BaseModel):
    email: str
    password: str
    role: str

class SubmissionCreate(BaseModel):
    manager_id: int
    total_invested: float
    total_gained: float
    total_lost: float
    notes: Optional[str] = None
    period: str
    sector: Optional[str] = None
    geography: Optional[str] = None
    liquidity_tier: Optional[str] = None
    vintage_year: Optional[int] = None
    committed_capital: Optional[float] = None
    called_capital: Optional[float] = None
    distributed_capital: Optional[float] = None
    residual_nav: Optional[float] = None
    entry_leverage: Optional[float] = None
    quartile_rank: Optional[str] = None
    quartile_source: Optional[str] = None

# ── SIGNUP ──
@router.post("/auth/signup")
def signup(data: SignupRequest, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.email == data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = User(
        name=data.name,
        email=data.email,
        password_hash=hash_password(data.password),
        role=data.role,
        asset_class=data.asset_class,
        institution_name=data.institution_name,
        institution_location=data.institution_location,
        job_title=data.job_title,
        phone_number=data.phone_number
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {
        "message": "Account created",
        "user_id": new_user.id,
        "name": new_user.name,
        "role": new_user.role,
        "asset_class": new_user.asset_class,
        "institution_name": new_user.institution_name
    }

# ── LOGIN ──
@router.post("/auth/login")
def login(data: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == data.email, User.role == data.role).first()
    if not user or not verify_password(data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    return {
        "message": "Login successful",
        "user_id": user.id,
        "name": user.name,
        "role": user.role,
        "asset_class": user.asset_class,
        "institution_name": user.institution_name
    }

# ── SUBMIT PORTFOLIO DATA ──
@router.post("/submissions/submit")
def submit_data(data: SubmissionCreate, db: Session = Depends(get_db)):
    existing = db.query(Submission).filter(
        Submission.manager_id == data.manager_id,
        Submission.period == data.period
    ).first()

    if existing:
        existing.total_invested = data.total_invested
        existing.total_gained = data.total_gained
        existing.total_lost = data.total_lost
        existing.notes = data.notes
        existing.sector = data.sector
        existing.geography = data.geography
        existing.liquidity_tier = data.liquidity_tier
        existing.vintage_year = data.vintage_year
        existing.committed_capital = data.committed_capital
        existing.called_capital = data.called_capital
        existing.distributed_capital = data.distributed_capital
        existing.residual_nav = data.residual_nav
        existing.entry_leverage = data.entry_leverage
        existing.quartile_rank = data.quartile_rank
        existing.quartile_source = data.quartile_source
        existing.submitted = True
        existing.submitted_at = datetime.utcnow()
        db.commit()
        return {"message": "Submission updated"}

    submission = Submission(
        manager_id=data.manager_id,
        total_invested=data.total_invested,
        total_gained=data.total_gained,
        total_lost=data.total_lost,
        notes=data.notes,
        period=data.period,
        sector=data.sector,
        geography=data.geography,
        liquidity_tier=data.liquidity_tier,
        vintage_year=data.vintage_year,
        committed_capital=data.committed_capital,
        called_capital=data.called_capital,
        distributed_capital=data.distributed_capital,
        residual_nav=data.residual_nav,
        entry_leverage=data.entry_leverage,
        quartile_rank=data.quartile_rank,
        quartile_source=data.quartile_source,
        submitted=True,
        submitted_at=datetime.utcnow()
    )
    db.add(submission)
    db.commit()
    return {"message": "Submission successful"}

# ── GET ALL SUBMISSIONS FOR A PERIOD (DIRECTOR VIEW) ──
@router.get("/submissions/{period}")
def get_submissions(period: str, institution: str = "", db: Session = Depends(get_db)):
    inst = institution.strip().lower()
    all_managers = db.query(User).filter(User.role == "manager").all()
    if inst:
        all_managers = [m for m in all_managers if (m.institution_name or "").strip().lower() == inst]
    manager_ids = {m.id for m in all_managers}
    submissions = db.query(Submission).filter(Submission.period == period).all()
    submissions = [s for s in submissions if s.manager_id in manager_ids]

    total_managers = len(all_managers)
    submitted_count = len(submissions)
    missing = total_managers - submitted_count

    manager_status = []
    for m in all_managers:
        sub = next((s for s in submissions if s.manager_id == m.id), None)
        manager_status.append({
            "manager_id": m.id,
            "name": m.name,
            "asset_class": m.asset_class,
            "submitted": sub is not None,
            "total_invested": sub.total_invested if sub else None,
            "total_gained": sub.total_gained if sub else None,
            "total_lost": sub.total_lost if sub else None,
            "submitted_at": sub.submitted_at if sub else None,
            "committed_capital": sub.committed_capital if sub else None,
            "called_capital": sub.called_capital if sub else None,
            "distributed_capital": sub.distributed_capital if sub else None,
            "residual_nav": sub.residual_nav if sub else None,
            "entry_leverage": sub.entry_leverage if sub else None,
            "quartile_rank": sub.quartile_rank if sub else None,
            "quartile_source": sub.quartile_source if sub else None
        })

    return {
        "period": period,
        "total_managers": total_managers,
        "submitted": submitted_count,
        "missing": missing,
        "complete": missing == 0,
        "managers": manager_status
    }

# ── GET MY SUBMISSION HISTORY (MANAGER VIEW) ──
@router.get("/submissions/manager/{manager_id}")
def get_manager_submissions(manager_id: int, db: Session = Depends(get_db)):
    submissions = db.query(Submission).filter(Submission.manager_id == manager_id).all()
    return {
        "submissions": [
            {
                "period": s.period,
                "total_invested": s.total_invested,
                "total_gained": s.total_gained,
                "total_lost": s.total_lost,
                "submitted_at": s.submitted_at,
                "committed_capital": s.committed_capital,
                "called_capital": s.called_capital,
                "distributed_capital": s.distributed_capital,
                "residual_nav": s.residual_nav,
                "entry_leverage": s.entry_leverage,
                "quartile_rank": s.quartile_rank,
                "quartile_source": s.quartile_source
            } for s in submissions
        ]
    }


def _period_sort_key(period: str):
    """Turns 'Q2 2026' into (2026, 2) so periods sort chronologically, not alphabetically."""
    try:
        q_part, year_part = period.split(" ")
        quarter = int(q_part.replace("Q", ""))
        year = int(year_part)
        return (year, quarter)
    except (ValueError, AttributeError):
        return (9999, 9)


# ── GET FULL SUBMISSION HISTORY ACROSS ALL PERIODS (FOR CHARTS) ──
@router.get("/submissions/history/all")
def get_all_submission_history(institution: str = "", db: Session = Depends(get_db)):
    inst = institution.strip().lower()
    submissions = db.query(Submission).filter(Submission.submitted == True).all()

    history = []
    for s in submissions:
        manager = db.query(User).filter(User.id == s.manager_id).first()
        if not manager:
            continue
        if inst and (manager.institution_name or "").strip().lower() != inst:
            continue
        net_change = s.total_gained - s.total_lost
        pct_return = (net_change / s.total_invested * 100) if s.total_invested else 0
        cumulative_value = s.total_invested + net_change

        history.append({
            "manager_id": s.manager_id,
            "manager_name": manager.name,
            "asset_class": manager.asset_class,
            "period": s.period,
            "total_invested": s.total_invested,
            "total_gained": s.total_gained,
            "total_lost": s.total_lost,
            "net_change": net_change,
            "pct_return": round(pct_return, 2),
            "cumulative_value": cumulative_value,
            "submitted_at": s.submitted_at.strftime("%b %d, %Y %H:%M") if s.submitted_at else "—",
            "sector": s.sector,
            "geography": s.geography,
            "liquidity_tier": s.liquidity_tier,
            "vintage_year": s.vintage_year,
            "committed_capital": s.committed_capital,
            "called_capital": s.called_capital,
            "distributed_capital": s.distributed_capital,
            "residual_nav": s.residual_nav,
            "entry_leverage": s.entry_leverage,
            "quartile_rank": s.quartile_rank,
            "quartile_source": s.quartile_source
        })

    history.sort(key=lambda h: _period_sort_key(h["period"]))

    return {"history": history}


# ── MARKET DATA PROXY (Alpha Vantage) ──
AV_KEY = "JPSR2HXKEFAJ9M3W"
AV_BASE = "https://www.alphavantage.co/query"

@router.get("/market/quote/{symbol}")
def get_quote(symbol: str):
    try:
        url = f"{AV_BASE}?function=GLOBAL_QUOTE&symbol={symbol}&apikey={AV_KEY}"
        with urllib.request.urlopen(url, timeout=8) as r:
            data = json_lib.loads(r.read())
        return data.get("Global Quote", {})
    except Exception as e:
        return {"error": str(e)}

@router.get("/market/treasury/{maturity}")
def get_treasury(maturity: str):
    try:
        url = f"{AV_BASE}?function=TREASURY_YIELD&interval=monthly&maturity={maturity}&apikey={AV_KEY}"
        with urllib.request.urlopen(url, timeout=8) as r:
            data = json_lib.loads(r.read())
        items = data.get("data", [])
        return {"value": items[0]["value"] if items else None}
    except Exception as e:
        return {"error": str(e)}

@router.get("/market/cpi")
def get_cpi():
    try:
        url = f"{AV_BASE}?function=CPI&interval=monthly&apikey={AV_KEY}"
        with urllib.request.urlopen(url, timeout=8) as r:
            data = json_lib.loads(r.read())
        items = data.get("data", [])
        return {"value": items[0]["value"] if items else None}
    except Exception as e:
        return {"error": str(e)}


# ── CIPHER AI CHAT ENDPOINT (Groq) ──
import httpx as _httpx
from fastapi import Request as _Request
from pydantic import BaseModel as _BaseModel
from typing import List as _List

class AIMessage(_BaseModel):
    messages: _List[dict]
    system: str

@router.post("/ai/chat")
async def ai_chat(data: AIMessage):
    try:
        # Prepend system message for Groq (uses messages format)
        msgs = [{"role": "system", "content": data.system}] + data.messages
        async with _httpx.AsyncClient() as client:
            res = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Authorization": "Bearer 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",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "llama-3.3-70b-versatile",
                    "max_tokens": 1000,
                    "messages": msgs
                },
                timeout=30
            )
        data_resp = res.json()
        # Return in Anthropic-style format so frontend works unchanged
        text = data_resp.get("choices", [{}])[0].get("message", {}).get("content", "")
        return {"content": [{"type": "text", "text": text}]}
    except Exception as e:
        return {"content": [{"type": "text", "text": f"Error: {str(e)}"}]}


# ── EXCEL EXPORT ──
from fastapi.responses import StreamingResponse
import io
from datetime import datetime

@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                      GradientFill)
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        NAVY   = "0A2472"
        SKY    = "4A90D9"
        SOFT   = "A8D4F5"
        WHITE  = "FFFFFF"
        GREEN  = "2E7D32"
        RED    = "C62828"
        LIGHT  = "F0F4FA"
        GOLD   = "F9A825"

        now = datetime.utcnow().strftime("%B %d, %Y")

        def hdr_font(size=11, bold=True, color=WHITE):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def body_font(size=10, bold=False, color="1A1A1A"):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def fill(color):
            return PatternFill("solid", fgColor=color)

        def border():
            s = Side(style="thin", color="DDDDDD")
            return Border(left=s, right=s, top=s, bottom=s)

        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def left():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)

        def add_header(ws, title, institution="CIPHER Platform", cols=10):
            # Row 1: CIPHER branding
            ws.merge_cells(f"A1:{get_column_letter(cols)}1")
            c1 = ws["A1"]
            c1.value = "CIPHER — CENTRAL INTELLIGENCE PLATFORM FOR HOLDINGS AND EXECUTIVE REPORTING"
            c1.font = Font(name="Arial", size=9, bold=True, color=SOFT)
            c1.fill = fill(NAVY)
            c1.alignment = center()
            ws.row_dimensions[1].height = 18

            # Row 2: Sheet title
            ws.merge_cells(f"A2:{get_column_letter(cols)}2")
            c2 = ws["A2"]
            c2.value = title
            c2.font = Font(name="Arial", size=14, bold=True, color=WHITE)
            c2.fill = fill(NAVY)
            c2.alignment = center()
            ws.row_dimensions[2].height = 28

            # Row 3: Meta info
            ws.merge_cells(f"A3:{get_column_letter(cols//2)}3")
            ws["A3"].value = f"Institution: {institution}"
            ws["A3"].font = Font(name="Arial", size=9, color="666666")
            ws["A3"].fill = fill(LIGHT)
            ws["A3"].alignment = left()

            ws.merge_cells(f"{get_column_letter(cols//2+1)}3:{get_column_letter(cols)}3")
            meta = ws[f"{get_column_letter(cols//2+1)}3"]
            meta.value = f"Report Date: {now}    |    CONFIDENTIAL — FOR INTERNAL USE ONLY"
            meta.font = Font(name="Arial", size=9, color="666666")
            meta.fill = fill(LIGHT)
            meta.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[3].height = 16

            # Row 4: blank spacer
            ws.row_dimensions[4].height = 8

            return 5  # next data row

        def add_section_title(ws, row, title, cols=10):
            ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
            cell = ws[f"A{row}"]
            cell.value = title
            cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            cell.fill = fill(SKY)
            cell.alignment = left()
            ws.row_dimensions[row].height = 18
            return row + 1

        def add_table_header(ws, row, headers, col_start=1):
            for i, h in enumerate(headers):
                cell = ws.cell(row=row, column=col_start+i, value=h)
                cell.font = hdr_font(size=9)
                cell.fill = fill(NAVY)
                cell.alignment = center()
                cell.border = border()
            ws.row_dimensions[row].height = 20
            return row + 1

        def add_table_row(ws, row, values, col_start=1, shade=False):
            bg = LIGHT if shade else WHITE
            for i, v in enumerate(values):
                cell = ws.cell(row=row, column=col_start+i, value=v)
                cell.font = body_font()
                cell.fill = fill(bg)
                cell.alignment = left() if i == 0 else center()
                cell.border = border()
            ws.row_dimensions[row].height = 16
            return row + 1

        def set_col_widths(ws, widths):
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = w

        # ─────────────────────────────────────────────────────────────
        # SHEET 1: DASHBOARD SUMMARY
        # ─────────────────────────────────────────────────────────────
        ws = wb.create_sheet("Dashboard Summary")
        row = add_header(ws, "INVESTMENT OPERATIONS DASHBOARD", cols=8)

        # Pull real data from DB
        all_managers = db.query(User).filter(User.role == "manager").all()
        submissions = db.query(Submission).filter(Submission.submitted == True).all()
        total_invested = sum(s.total_invested for s in submissions)
        total_gained   = sum(s.total_gained   for s in submissions)
        total_lost     = sum(s.total_lost     for s in submissions)
        net_change     = total_gained - total_lost
        submitted_count = len(set(s.manager_id for s in submissions))

        row = add_section_title(ws, row, "PORTFOLIO SUMMARY", cols=8)
        row = add_table_header(ws, row, ["METRIC","VALUE","NOTES"], col_start=1)
        summary_rows = [
            ("Total Invested",  f"${total_invested:,.0f}",  "Across all managers"),
            ("Total Gained",    f"${total_gained:,.0f}",    "This period"),
            ("Total Lost",      f"${total_lost:,.0f}",      "This period"),
            ("Net Change",      f"${net_change:+,.0f}",     "Gained minus lost"),
            ("Managers Filed",  f"{submitted_count} / {len(all_managers)}", "Submission rate"),
        ]
        for i, r in enumerate(summary_rows):
            row = add_table_row(ws, row, r, shade=i%2==0)

        row += 1
        row = add_section_title(ws, row, "MANAGER SUBMISSIONS", cols=8)
        row = add_table_header(ws, row, ["MANAGER","ASSET CLASS","INVESTED","GAINED","LOST","NET CHANGE","STATUS","SUBMITTED AT"])
        for i, m in enumerate(all_managers):
            sub = next((s for s in submissions if s.manager_id == m.id), None)
            row = add_table_row(ws, row, [
                m.name,
                m.asset_class or "—",
                f"${sub.total_invested:,.0f}" if sub else "—",
                f"${sub.total_gained:,.0f}"   if sub else "—",
                f"${sub.total_lost:,.0f}"     if sub else "—",
                f"${(sub.total_gained - sub.total_lost):+,.0f}" if sub else "—",
                "Submitted" if sub else "Missing",
                sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub and sub.submitted_at else "—",
            ], shade=i%2==0)
            if sub and not sub.submitted_at == None:
                status_cell = ws.cell(row=row-1, column=7)
                status_cell.font = Font(name="Arial", size=10, bold=True,
                    color=GREEN if sub else RED)

        set_col_widths(ws, [22, 16, 14, 14, 14, 14, 12, 18])

        # ─────────────────────────────────────────────────────────────
        # SHEET 2: PORTFOLIO PERFORMANCE
        # ─────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Portfolio Performance")
        row = add_header(ws2, "PORTFOLIO PERFORMANCE", cols=9)

        row = add_section_title(ws2, row, "TIME-PERIOD RETURNS", cols=9)
        row = add_table_header(ws2, row, ["ASSET CLASS","CCY","QTD","YTD","1 YEAR","3 YEAR","5 YEAR","SINCE INCEPTION","INCEPTION DATE"])
        perf_data = [
            ("Total Portfolio", "USD", "+2.1%", "+8.4%", "+11.2%", "+11.8%", "+9.4%",  "+94.2%",  "Jan 2015"),
            ("Public Equity",   "USD", "+2.8%", "+11.2%","+14.1%", "+13.4%", "+11.2%", "+112.4%", "Jan 2015"),
            ("Private Equity",  "USD", "+1.9%", "+9.8%", "+12.4%", "+14.2%", "+12.8%", "+98.6%",  "Mar 2016"),
            ("Fixed Income",    "USD", "+0.8%", "+4.1%", "+5.2%",  "+4.8%",  "+3.9%",  "+42.1%",  "Jan 2015"),
            ("Real Assets",     "USD", "+1.4%", "+6.7%", "+8.9%",  "+9.2%",  "+8.1%",  "+74.8%",  "Jun 2016"),
            ("Operations",      "USD", "-0.3%", "-1.2%", "-0.8%",  "+2.1%",  "+1.8%",  "+18.4%",  "Jan 2015"),
            ("S&P 500",         "USD", "+1.4%", "+6.2%", "+9.4%",  "+10.2%", "+8.8%",  "+86.4%",  "Benchmark"),
            ("MSCI World",      "USD", "+1.2%", "+5.8%", "+8.6%",  "+9.1%",  "+7.9%",  "+74.2%",  "Benchmark"),
        ]
        for i, r in enumerate(perf_data):
            row = add_table_row(ws2, row, r, shade=i%2==0)
            # Color returns
            for col in range(3, 9):
                cell = ws2.cell(row=row-1, column=col)
                if cell.value and str(cell.value).startswith("+"):
                    cell.font = Font(name="Arial", size=10, color=GREEN)
                elif cell.value and str(cell.value).startswith("-"):
                    cell.font = Font(name="Arial", size=10, color=RED)

        row += 1
        row = add_section_title(ws2, row, "BENCHMARK COMPARISON", cols=9)
        row = add_table_header(ws2, row, ["BENCHMARK","YTD RETURN","PORTFOLIO vs BENCHMARK","ALPHA"])
        bench_data = [
            ("Portfolio",      "+8.4%",  "—",      "—"),
            ("S&P 500",        "+6.2%",  "+2.2%",  "+2.2%"),
            ("MSCI World",     "+5.8%",  "+2.6%",  "+2.6%"),
            ("Bloomberg US Agg","+2.1%", "+6.3%",  "+6.3%"),
        ]
        for i, r in enumerate(bench_data):
            row = add_table_row(ws2, row, r, shade=i%2==0)

        set_col_widths(ws2, [20, 8, 10, 10, 10, 10, 10, 16, 14])

        # ─────────────────────────────────────────────────────────────
        # SHEET 3: PRIVATE MARKETS
        # ─────────────────────────────────────────────────────────────
        ws3 = wb.create_sheet("Private Markets")
        row = add_header(ws3, "PRIVATE MARKETS — CAPITAL ACCOUNTS", cols=10)

        # DPI/RVPI/TVPI summary row
        row = add_section_title(ws3, row, "FUND MULTIPLE SUMMARY", cols=10)
        row = add_table_header(ws3, row, ["METRIC","VALUE","FORMULA","NOTES"])
        multiples = [
            ("DPI",  "0.37x", "Distributions / Capital Called", "$42.1M / $112.4M"),
            ("RVPI", "1.11x", "NAV / Capital Called",           "$124.6M / $112.4M"),
            ("TVPI", "1.48x", "(Distributions + NAV) / Called", "($42.1M + $124.6M) / $112.4M"),
            ("Net IRR (Composite)", "14.2%", "Since inception", "Weighted average all funds"),
        ]
        for i, r in enumerate(multiples):
            row = add_table_row(ws3, row, r, shade=i%2==0)

        row += 1
        row = add_section_title(ws3, row, "CAPITAL ACCOUNT BY FUND", cols=10)
        row = add_table_header(ws3, row, ["FUND","VINTAGE","COMMITTED","CALLED","UNCALLED","DISTRIBUTED","NAV","DPI","TVPI","NET IRR"])
        fund_data = [
            ("Apex Growth III",    "2021", "$12.0M","$9.6M", "$2.4M", "$2.1M",  "$13.6M","0.22x","1.63x","18.4%"),
            ("Meridian PE II",     "2020", "$8.5M", "$8.5M", "—",     "$6.2M",  "$11.1M","0.73x","2.04x","12.1%"),
            ("Blueridge Ventures", "2022", "$6.0M", "$3.2M", "$2.8M", "$0.4M",  "$2.8M", "0.13x","0.88x","-2.4%"),
            ("Horizon Real Assets","2019", "$14.0M","$14.0M","—",     "$11.8M", "$23.5M","0.84x","2.52x","16.7%"),
            ("ClearPath Infra I",  "2023", "$7.7M", "$2.1M", "$5.6M", "—",      "$2.1M", "0.00x","1.00x","—"),
        ]
        for i, r in enumerate(fund_data):
            row = add_table_row(ws3, row, r, shade=i%2==0)
            irr_cell = ws3.cell(row=row-1, column=10)
            if irr_cell.value and str(irr_cell.value).startswith("-"):
                irr_cell.font = Font(name="Arial", size=10, color=RED)
            elif irr_cell.value and irr_cell.value != "—":
                irr_cell.font = Font(name="Arial", size=10, color=GREEN)

        set_col_widths(ws3, [22, 10, 12, 12, 12, 14, 12, 10, 10, 10])

        # ─────────────────────────────────────────────────────────────
        # SHEET 4: RISK METRICS
        # ─────────────────────────────────────────────────────────────
        ws4 = wb.create_sheet("Risk Metrics")
        row = add_header(ws4, "RISK METRICS", cols=7)

        row = add_section_title(ws4, row, "RISK SUMMARY", cols=7)
        row = add_table_header(ws4, row, ["METRIC","VALUE","BENCHMARK","NOTES"])
        risk_summary = [
            ("Sharpe Ratio (3Y)",       "1.42",   "1.10",   "Risk-adjusted return"),
            ("Standard Deviation (3Y)", "8.2%",   "12.4%",  "Annualized volatility"),
            ("Max Drawdown",            "-14.8%", "-18.2%", "Peak to trough since inception"),
            ("Value at Risk (95%)",     "-2.1%",  "-3.4%",  "Monthly VaR estimate"),
            ("Beta vs S&P 500",         "0.61",   "1.00",   "Market correlation"),
        ]
        for i, r in enumerate(risk_summary):
            row = add_table_row(ws4, row, r, shade=i%2==0)

        row += 1
        row = add_section_title(ws4, row, "RISK BY ASSET CLASS", cols=7)
        row = add_table_header(ws4, row, ["ASSET CLASS","3Y RETURN","STD DEV","SHARPE","MAX DRAWDOWN","VAR (95%)","BETA"])
        risk_data = [
            ("Public Equity",  "+13.4%","12.1%","1.11","-18.2%","-3.2%","0.84"),
            ("Private Equity", "+14.2%","6.8%", "2.09","-8.4%", "-1.4%","0.42"),
            ("Fixed Income",   "+4.8%", "3.2%", "1.50","-4.1%", "-0.8%","-0.12"),
            ("Real Assets",    "+9.2%", "7.4%", "1.24","-11.2%","-1.9%","0.31"),
            ("Operations",     "+2.1%", "4.8%", "0.44","-9.8%", "-1.2%","0.18"),
            ("Total Portfolio","+11.8%","8.2%", "1.42","-14.8%","-2.1%","0.61"),
        ]
        for i, r in enumerate(risk_data):
            row = add_table_row(ws4, row, r, shade=i%2==0)

        set_col_widths(ws4, [20, 12, 12, 12, 16, 12, 10])

        # ─────────────────────────────────────────────────────────────
        # SHEET 5: EXPOSURE ANALYSIS
        # ─────────────────────────────────────────────────────────────
        ws5 = wb.create_sheet("Exposure Analysis")
        row = add_header(ws5, "EXPOSURE ANALYSIS", cols=6)

        row = add_section_title(ws5, row, "SECTOR EXPOSURE", cols=6)
        row = add_table_header(ws5, row, ["SECTOR","EXPOSURE","% OF PORTFOLIO","LONG","SHORT","NET"])
        sector_data = [
            ("Healthcare",     "$43.6M","35.0%","$43.6M","—",     "$43.6M"),
            ("Technology",     "$24.9M","20.0%","$24.9M","$2.1M", "$22.8M"),
            ("Real Estate",    "$18.7M","15.0%","$18.7M","—",     "$18.7M"),
            ("Infrastructure", "$12.5M","10.0%","$12.5M","—",     "$12.5M"),
            ("Other",          "$24.9M","20.0%","$24.9M","$1.2M", "$23.7M"),
        ]
        for i, r in enumerate(sector_data):
            row = add_table_row(ws5, row, r, shade=i%2==0)

        row += 1
        row = add_section_title(ws5, row, "GEOGRAPHY EXPOSURE", cols=6)
        row = add_table_header(ws5, row, ["REGION","EXPOSURE","% OF PORTFOLIO","CURRENCY","FX RATE","USD VALUE"])
        geo_data = [
            ("North America", "$77.7M","62.4%","USD","1.00","$77.7M"),
            ("Europe",        "$22.4M","18.0%","EUR","1.08","$22.4M"),
            ("Asia Pacific",  "$14.9M","12.0%","JPY","0.0067","$14.9M"),
            ("Latin America", "$6.2M", "5.0%", "USD","1.00","$6.2M"),
            ("Other",         "$3.4M", "2.6%", "Various","—","$3.4M"),
        ]
        for i, r in enumerate(geo_data):
            row = add_table_row(ws5, row, r, shade=i%2==0)

        set_col_widths(ws5, [20, 14, 16, 12, 10, 14])

        # ─────────────────────────────────────────────────────────────
        # SHEET 6: CASH MANAGEMENT
        # ─────────────────────────────────────────────────────────────
        ws6 = wb.create_sheet("Cash Management")
        row = add_header(ws6, "CASH MANAGEMENT — CALLS & DISTRIBUTIONS", cols=7)

        row = add_section_title(ws6, row, "UPCOMING CAPITAL CALLS", cols=7)
        row = add_table_header(ws6, row, ["FUND","CALL DATE","AMOUNT","PURPOSE","UNCALLED REMAINING","STATUS","PRIORITY"])
        calls_data = [
            ("Apex Growth III",    "Jul 15, 2026","$1.2M","Follow-on investment","$1.2M after","Due Soon","HIGH"),
            ("ClearPath Infra I",  "Jul 28, 2026","$3.0M","Construction milestone","$2.6M after","Due Soon","HIGH"),
            ("Blueridge Ventures", "Sep 1, 2026", "$1.4M","Series B investment",  "$1.4M after","Upcoming","MEDIUM"),
            ("ClearPath Infra I",  "Oct 15, 2026","$1.2M","Phase 2 construction", "$1.4M after","Upcoming","MEDIUM"),
            ("Apex Growth III",    "Dec 1, 2026", "$1.2M","New portfolio company", "—",          "Upcoming","LOW"),
        ]
        for i, r in enumerate(calls_data):
            row = add_table_row(ws6, row, r, shade=i%2==0)

        row += 1
        row = add_section_title(ws6, row, "DISTRIBUTION HISTORY", cols=7)
        row = add_table_header(ws6, row, ["DATE","FUND","TYPE","AMOUNT","CUMULATIVE DPI","NOTES"])
        dist_data = [
            ("Jun 2026","Horizon Real Assets","Income",         "$2.4M","0.84x","Quarterly income distribution"),
            ("Mar 2026","Meridian PE II",      "Realized Gain",  "$1.8M","0.73x","Partial exit — portfolio company"),
            ("Dec 2025","Horizon Real Assets", "Income",         "$2.1M","0.72x","Quarterly income distribution"),
            ("Sep 2025","Meridian PE II",      "Recapitalization","$2.6M","0.52x","Recap of portfolio company"),
            ("Jun 2025","Apex Growth III",     "Partial Sale",   "$2.1M","0.22x","Secondary sale of position"),
        ]
        for i, r in enumerate(dist_data):
            row = add_table_row(ws6, row, r, shade=i%2==0)

        set_col_widths(ws6, [22, 14, 16, 12, 16, 10, 10])

        # ─────────────────────────────────────────────────────────────
        # SHEET 7: COMPLIANCE & AUDIT
        # ─────────────────────────────────────────────────────────────
        ws7 = wb.create_sheet("Compliance & Audit")
        row = add_header(ws7, "COMPLIANCE — IPS TRACKING & AUDIT LOG", cols=8)

        row = add_section_title(ws7, row, "IPS POLICY TRACKING", cols=8)
        row = add_table_header(ws7, row, ["ASSET CLASS","POLICY MIN","TARGET","POLICY MAX","ACTUAL","VARIANCE","IN POLICY","NOTES"])
        ips_data = [
            ("Public Equity",  "25%","35%","45%","33.8%","-1.2%","YES","Within range"),
            ("Private Equity", "15%","25%","35%","25.3%","+0.3%","YES","Within range"),
            ("Fixed Income",   "15%","20%","30%","19.9%","-0.1%","YES","Within range"),
            ("Real Assets",    "10%","15%","20%","14.6%","-0.4%","YES","Within range"),
            ("Operations",     "5%", "5%", "10%","6.4%", "+1.4%","YES","Within range"),
        ]
        for i, r in enumerate(ips_data):
            row = add_table_row(ws7, row, r, shade=i%2==0)
            in_policy_cell = ws7.cell(row=row-1, column=7)
            in_policy_cell.font = Font(name="Arial", size=10, bold=True,
                color=GREEN if r[6]=="YES" else RED)

        row += 1
        row = add_section_title(ws7, row, "AUDIT LOG", cols=8)
        row = add_table_header(ws7, row, ["TIMESTAMP","USER","ACTION","DETAILS","STATUS","IP ADDRESS"])
        audit_data = [
            ("Jul 3, 2026 14:22","Sarah Johnson","SUBMISSION",      "Q2 2026 — Private Equity — $31.5M","OK",      "192.168.1.42"),
            ("Jul 3, 2026 11:14","Mike Chen",    "SUBMISSION",      "Q2 2026 — Real Assets — $18.2M",  "OK",      "192.168.1.38"),
            ("Jul 2, 2026 16:48","Priya Patel",  "SUBMISSION EDIT", "Q1 2026 — values modified",        "FLAGGED", "192.168.1.51"),
            ("Jul 2, 2026 09:31","Director",     "LOGIN",           "Investment manager portal",         "OK",      "192.168.1.10"),
            ("Jul 1, 2026 17:22","Alex Torres",  "MISSED DEADLINE", "Q2 2026 — no submission",          "FLAGGED", "—"),
        ]
        for i, r in enumerate(audit_data):
            row = add_table_row(ws7, row, r, shade=i%2==0)
            status_cell = ws7.cell(row=row-1, column=5)
            status_cell.font = Font(name="Arial", size=10, bold=True,
                color=RED if r[4]=="FLAGGED" else GREEN)

        set_col_widths(ws7, [22, 16, 18, 36, 10, 14])

        # ─────────────────────────────────────────────────────────────
        # SAVE & STREAM
        # ─────────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"CIPHER_Export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return {"error": str(e)}


# ── EMAIL VERIFICATION (2FA) — via Brevo HTTPS API ──
# (SMTP is blocked on Railway's free plan, so we send over HTTPS instead)
BREVO_API_KEY = os.environ.get("BREVO_API_KEY", "")
SENDER_EMAIL  = os.environ.get("SENDER_EMAIL", "noreplycipher@gmail.com")

verification_codes = {}  # {user_id: "123456"}

def send_verification_email(to_email: str, code: str):
    payload = json_lib.dumps({
        "sender": {"email": SENDER_EMAIL, "name": "CIPHER"},
        "to": [{"email": to_email}],
        "subject": "Your CIPHER Verification Code",
        "textContent": f"Your CIPHER verification code is: {code}\n\nIf you didn't request this, you can ignore this email."
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email",
        data=payload,
        headers={
            "accept": "application/json",
            "api-key": BREVO_API_KEY,
            "content-type": "application/json"
        },
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        resp.read()

class SendCodeRequest(BaseModel):
    user_id: int
    email: str

class VerifyCodeRequest(BaseModel):
    user_id: int
    code: str

@router.post("/auth/send-code")
def send_code(data: SendCodeRequest):
    code = str(random.randint(100000, 999999))
    verification_codes[data.user_id] = code
    if not BREVO_API_KEY:
        print(f"[CIPHER 2FA - EMAIL API NOT CONFIGURED] Code for {data.email}: {code}")
        return {"message": "Code generated (email API not configured — check Railway logs for the code)"}
    try:
        send_verification_email(data.email, code)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not send email: {str(e)}")
    return {"message": "Code sent"}

@router.post("/auth/verify-code")
def verify_code(data: VerifyCodeRequest):
    expected = verification_codes.get(data.user_id)
    if not expected or expected != data.code:
        raise HTTPException(status_code=401, detail="Incorrect or expired code")
    del verification_codes[data.user_id]
    return {"message": "Verified"}
